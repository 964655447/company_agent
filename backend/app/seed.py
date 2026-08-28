"""花名册 seed：把 花名册(1).xlsx 导入 employees 表（密码 bcrypt 哈希后入库）。

用法：
  python -m app.seed          # 空库时导入；--force 覆盖重导
"""
import sys

import openpyxl
from sqlalchemy import select

from .config import ROSTER_XLSX
from .database import SessionLocal, Base, engine
from .models import Employee
from .security import hash_password


def seed_employees(force: bool = False) -> int:
    Base.metadata.create_all(bind=engine)
    with SessionLocal() as db:
        existing = db.scalar(select(Employee).limit(1))
        if existing and not force:
            print(f"[seed] 库中已有 {db.query(Employee).count()} 名员工，跳过。（--force 可覆盖重导）")
            return 0
        if existing and force:
            db.query(Employee).delete()
            db.commit()

        try:
            wb = openpyxl.load_workbook(ROSTER_XLSX)
        except FileNotFoundError:
            print(f"[seed] 找不到花名册文件：{ROSTER_XLSX}，跳过导入")
            return 0

        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        header, data = rows[0], rows[1:]
        # 列：NO / Name / ID / Login Password / Permissions / Position / Department
        count = 0
        for r in data:
            if not r or not r[1]:
                continue
            no_, name, emp_id, pwd, perms, position, dept = r[:7]
            permissions = [p.strip() for p in str(perms or "").split("、") if p.strip()]
            db.add(Employee(
                no=int(no_), emp_id=int(emp_id), name=str(name).strip(),
                password_hash=hash_password(str(pwd)),
                permissions=__import__("json").dumps(permissions, ensure_ascii=False),
                position=str(position).strip(), department=str(dept).strip(),
            ))
            count += 1
        db.commit()
        print(f"[seed] 已导入 {count} 名员工（密码已哈希入库）")
        return count


if __name__ == "__main__":
    seed_employees(force="--force" in sys.argv)
