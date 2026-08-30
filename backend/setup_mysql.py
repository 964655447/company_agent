"""一次性初始化本地 MySQL：建库建表（schema 来自 ../contracts/database.sql）。

同事 clone 仓库后的标准流程（也可由启动器 launcher 自动调用，无需手动跑）：
  1. cd backend
  2. cp ../.env.example .env        （或首次启动启动器时网页填写数据库信息）
  3. pip install -r requirements.txt
  4. python setup_mysql.py                     # 自动建库 + 6 张表（幂等，可重复跑）
  5.（可选）python setup_mysql.py --seed-demo         # 程序化生成演示数据（员工/考勤/薪资/报销/考核查询日志/考核成绩，幂等，已非空表跳过）
                                                  # 加 --reset 则先清空 6 张表再重灌
  6.（可选）python setup_mysql.py --seed-employees   # 需本机有花名册.xlsx
                                                  # 再设环境变量 SEED_ATTENDANCE=1 可生成演示考勤

注意：
  - 本文件只负责「结构」。员工/考勤等数据由各人自己的花名册.xlsx 灌入，不随仓库分发。
  - ensure_database() 会 CREATE DATABASE IF NOT EXISTS 并建表；表已存在则跳过 DDL（不覆盖数据）。
    要在本地启用新结构（如工资表 v2），请执行 contracts/migrate_salary_v2.sql。
"""
import os
import re
import sys
from pathlib import Path

try:
    import pymysql
except ImportError:
    raise SystemExit(
        "[依赖缺失] 当前 Python 环境未安装 pymysql。\n"
        "请执行以下命令安装：\n"
        "  pip install pymysql\n"
        "或（如果用的是 Anaconda）：\n"
        "  conda install pymysql -c conda-forge\n"
        "安装后重新运行本脚本即可。"
    )
import random
import json
from datetime import date, datetime, timedelta

BASE_DIR = Path(__file__).resolve().parent
ROOT_DIR = BASE_DIR.parent
SQL_PATH = ROOT_DIR / "contracts" / "database.sql"
ASSESSMENT_SEED_SQL = ROOT_DIR / "contracts" / "assessment_stats_seed.sql"


# ---------- 轻量读取 .env（不依赖 python-dotenv）----------
def load_dotenv() -> None:
    p = BASE_DIR / ".env"
    if not p.exists():
        return
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        k, v = k.strip(), v.strip().strip('"').strip("'")
        if k and k not in os.environ:
            os.environ[k] = v


load_dotenv()


# ---------- 解析 DB_URL ----------
def parse_db_url(url: str) -> dict:
    # 兼容 mysql+pymysql:// 与 mysql:// 两种写法
    m = re.match(r"mysql\+?pymysql?://([^:]+):([^@]+)@([^:/]+):?(\d+)?/([^?]+)", url)
    if not m:
        raise SystemExit(
            f"[配置错误] 无法解析 DB_URL：{url}\n"
            f"示例：mysql+pymysql://root:密码@127.0.0.1:3306/company_agent"
        )
    user, pwd, host, port, db = m.groups()
    return dict(host=host, port=int(port or 3306), user=user, password=pwd, db=db)


DB_URL = os.environ.get("DB_URL")
if not DB_URL:
    raise SystemExit(
        "[配置错误] 未设置 DB_URL。\n"
        "请在本文件同目录的 .env 中设置（可复制 .env.example）：\n"
        "  DB_URL=mysql+pymysql://<用户>:<密码>@<主机>:<端口>/<库名>\n"
        "例如：DB_URL=mysql+pymysql://root:你的密码@127.0.0.1:3306/company_agent\n"
        "（首次启动启动器时也会在网页向导中填写并写入 .env，无需手动设）"
    )
MYSQL = parse_db_url(DB_URL)
DB_NAME = MYSQL.pop("db")
ROSTER_XLSX = os.environ.get("ROSTER_XLSX", "")
SEED_ATTENDANCE = os.environ.get("SEED_ATTENDANCE", "") == "1"


def connect(**kw):
    return pymysql.connect(charset="utf8mb4", cursorclass=pymysql.cursors.DictCursor, **kw)


def run_schema(conn):
    """执行 database.sql 中的所有 DDL 语句。

    每条 CREATE TABLE 均带 IF NOT EXISTS，已存在的表自动跳过；
    缺失的表会被创建。因此无论库处于什么状态（空/部分建表/完整），
    调用本函数后都能保证 6 张表全部存在。
    """
    sql = SQL_PATH.read_text(encoding="utf-8")
    stmts, buf = [], ""
    for line in sql.splitlines():
        if line.strip().startswith("--"):
            continue
        buf += line + "\n"
        if line.strip().endswith(";"):
            stmts.append(buf.strip())
            buf = ""
    created, skipped, errors = 0, 0, []
    with conn.cursor() as cur:
        for s in stmts:
            try:
                cur.execute(s)
                created += 1
            except pymysql.err.OperationalError as e:
                # IF NOT EXISTS 已覆盖大部分情况；这里兜底捕获并记录
                if e.args[0] == 1050:  # "Table already exists"
                    skipped += 1
                else:
                    errors.append(str(e))
    conn.commit()
    msg = f"[schema] DDL 执行完毕：{created} 条成功"
    if skipped:
        msg += f"，{skipped} 条跳过（已存在）"
    if errors:
        msg += f"，{len(errors)} 条异常：{'; '.join(errors[:3])}"
    print(msg)


def ensure_database(conn=None):
    """确保数据库与表存在（幂等）。返回结果 dict。

    - 库不存在则 CREATE DATABASE IF NOT EXISTS；
    - 已存在则跳过建库；
    - 进入库后调用 run_schema 建表（已有表跳过 DDL）。
    可被启动器 subprocess 调用，也可直接 import 使用。
    """
    own = conn is None
    if own:
        conn = connect(**{k: v for k, v in MYSQL.items()})
    try:
        with conn.cursor() as cur:
            cur.execute(
                f"CREATE DATABASE IF NOT EXISTS `{DB_NAME}` "
                f"CHARACTER SET utf8mb4 COLLATE utf8mb4_general_ci"
            )
        conn.select_db(DB_NAME)
        run_schema(conn)
        return {"ok": True, "msg": f"数据库 {DB_NAME} 已就绪（建库/建表完成）"}
    except Exception as e:
        return {"ok": False, "msg": f"数据库初始化失败: {e}"}
    finally:
        if own:
            conn.close()


def seed_employees(conn):
    if not ROSTER_XLSX or not Path(ROSTER_XLSX).exists():
        print("[employees] 未设置 ROSTER_XLSX 或文件不存在，跳过导入"
              "（同事用各自花名册；管理者也可直接 INSERT employees 表）")
        return
    import json
    import bcrypt
    import openpyxl

    wb = openpyxl.load_workbook(ROSTER_XLSX)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header, data = rows[0], rows[1:]
    print("[employees] 列头:", header)

    with conn.cursor() as cur:
        cur.execute(f"SELECT COUNT(*) AS c FROM {DB_NAME}.employees")
        if cur.fetchone()["c"] > 0:
            print("[employees] 已有数据，跳过导入（如需重导请先清空表）")
            return
        ins = (
            "INSERT INTO employees (no, employee_id, name, password_hash, permissions, position, department) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s)"
        )
        for r in data:
            if not r or not r[1]:
                continue
            no_, name, emp_id, pwd, perms, position, dept = r[:7]
            permissions = [p.strip() for p in str(perms or "").split("、") if p.strip()]
            h = bcrypt.hashpw(str(pwd).encode(), bcrypt.gensalt()).decode()
            cur.execute(ins, (int(no_), int(emp_id), str(name).strip(), h,
                              json.dumps(permissions, ensure_ascii=False),
                              str(position).strip(), str(dept).strip()))
        conn.commit()
        print(f"[employees] 已导入 {len(data)} 名员工（bcrypt 哈希）")


def _gen_attendance_month(conn, cur, emp_ids, year, month, last_day):
    """为指定月份（含 last_day 当天）生成每位员工的上下班打卡，返回插入条数。"""
    workdays = []
    d = date(year, month, 1)
    while d <= last_day:
        if d.weekday() < 5:  # 周一~周五
            workdays.append(d)
        d += timedelta(days=1)
    ins = ("INSERT INTO attendance (employee_id, checkin_time, type, is_late, work_date) "
           "VALUES (%s,%s,%s,%s,%s)")
    n = 0
    for eid in emp_ids:
        for wd in workdays:
            late = random.random() < 0.12
            minute = random.randint(3, 42) if late else random.randint(0, 55)
            cin = datetime(wd.year, wd.month, wd.day, 9, minute)
            cur.execute(ins, (eid, cin, "clock_in", late, wd))
            cout = datetime(wd.year, wd.month, wd.day, 18, random.randint(0, 45))
            cur.execute(ins, (eid, cout, "clock_out", False, wd))
            n += 2
    return n


def _recent_months(today, months_back=2):
    """返回 [当月, 前1月, 前2月, ...] 的 (year, month) 列表（含当月共 months_back+1 个月）。"""
    months = [(today.year, today.month)]
    y, m = today.year, today.month
    for _ in range(months_back):
        m -= 1
        if m == 0:
            m = 12
            y -= 1
        months.append((y, m))
    return months


def _last_day_of(yy, mm):
    nxt = date(yy + 1, 1, 1) if mm == 12 else date(yy, mm + 1, 1)
    return nxt - timedelta(days=1)


def gen_attendance(conn):
    if not SEED_ATTENDANCE:
        print("[attendance] 未开启 SEED_ATTENDANCE，跳过（演示数据，同事一般不需）")
        return

    with conn.cursor() as cur:
        cur.execute(f"SELECT employee_id FROM {DB_NAME}.employees ORDER BY employee_id")
        emp_ids = [row["employee_id"] for row in cur.fetchall()]
        cur.execute(f"SELECT COUNT(*) AS c FROM {DB_NAME}.attendance")
        if cur.fetchone()["c"] > 0:
            print("[attendance] 已有数据，跳过生成")
            return
        today = date.today()
        total = 0
        for (yy, mm) in _recent_months(today, months_back=2):
            last_day = today if (yy == today.year and mm == today.month) else _last_day_of(yy, mm)
            total += _gen_attendance_month(conn, cur, emp_ids, yy, mm, last_day)
        conn.commit()
        print(f"[attendance] 已生成 {total} 条记录（{len(emp_ids)} 人 × 近 3 个月）")



# ------------------------------------------------------------
# 一键演示数据：程序化生成（不依赖外部 seed.sql / 花名册.xlsx）。
# 用法：
#   python setup_mysql.py --seed-demo            # 对空表造数（各表自带幂等守卫，已非空跳过）
#   python setup_mysql.py --seed-demo --reset    # 清空 5 张表后重新造数
# 说明：employees 为空时插入 27 名演示员工（工号 220401+，密码统一 123456）；
#       其余表按现有员工生成考勤/薪资/报销/考核查询日志/考核成绩演示数据。全部幂等，重复跑不重复造。
# ------------------------------------------------------------
# 演示花名册（name, department, position, is_admin）。仅在 employees 为空时插入。
DEMO_ROSTER = [
    ("冯霞", "总经办", "总经理", True),
    ("蒋博泺", "技术部", "技术经理", False),
    ("许梦", "商务部", "营销专员", False),
    ("张伟", "技术部", "开发工程师", False),
    ("李娜", "人事部", "人事专员", False),
    ("王芳", "财务部", "会计", False),
    ("刘洋", "运营部", "运营专员", False),
    ("陈静", "客服部", "客服专员", False),
    ("杨磊", "车队", "车队长", False),
    ("赵敏", "商务部", "营销经理", False),
    ("孙强", "技术部", "开发工程师", False),
    ("周婷", "人事部", "人事经理", False),
    ("吴军", "财务部", "出纳", False),
    ("郑昊", "运营部", "运营经理", False),
    ("黄丽", "客服部", "客服经理", False),
    ("马涛", "车队", "司机", False),
    ("朱琳", "财务部", "财务总监", False),
    ("胡斌", "技术部", "测试工程师", False),
    ("郭倩", "商务部", "商务专员", False),
    ("林峰", "运营部", "调度员", False),
    ("何敏", "客服部", "客服专员", False),
    ("高翔", "车队", "司机", False),
    ("罗静", "人事部", "招聘专员", False),
    ("梁宇", "技术部", "运维工程师", False),
    ("宋佳", "商务部", "营销专员", False),
    ("唐磊", "财务部", "审核", False),
    ("韩雪", "总经办", "总经理助理", False),
]
ALL_DEPTS = ["总经办", "商务部", "技术部", "人事部", "财务部", "运营部", "客服部", "车队"]
DEMO_SALARY_MAP = {
    "总经理": 20000, "总经理助理": 12000, "商务总监": 15000, "运营总监": 15000,
    "财务总监": 15000, "营销经理": 10000, "调度经理": 10000, "运营经理": 10000,
    "客服经理": 10000, "总车队长": 12000, "车队长": 8000, "营销专员": 6000,
    "调度专员": 6000, "运营专员": 6000, "客服专员": 6000, "车队文员": 5000,
    "会计": 8000, "出纳": 6000, "审核": 7000, "技术经理": 14000, "开发工程师": 9000,
    "测试工程师": 9000, "运维工程师": 9000, "人事经理": 9000, "人事专员": 6000,
    "招聘专员": 6000, "商务专员": 6000, "调度员": 6000, "司机": 5000,
}
DEMO_SALARY_DEFAULT = 6000


def gen_demo_dataset(conn):
    """一键生成自包含演示数据（employees 为空时插入演示员工，其余表按现有员工造数）。幂等。"""
    import bcrypt
    with conn.cursor() as cur:
        # 1) employees（仅当空表时插入演示员工）
        cur.execute(f"SELECT COUNT(*) c FROM {DB_NAME}.employees")
        if cur.fetchone()["c"] == 0:
            ins = ("INSERT INTO employees (no, employee_id, name, password_hash, "
                   "permissions, position, department) VALUES (%s,%s,%s,%s,%s,%s,%s)")
            h = bcrypt.hashpw("123456".encode(), bcrypt.gensalt()).decode()
            for i, (name, dept, pos, admin) in enumerate(DEMO_ROSTER):
                perms = json.dumps(ALL_DEPTS if admin else [dept], ensure_ascii=False)
                cur.execute(ins, (i + 1, 220401 + i, name, h, perms, pos, dept))
            conn.commit()
            print(f"[demo] 已插入 {len(DEMO_ROSTER)} 名演示员工（工号 220401+，密码统一 123456）")
        else:
            print("[demo] employees 已有数据，跳过")

        cur.execute(
            f"SELECT employee_id, name, position FROM {DB_NAME}.employees ORDER BY employee_id"
        )
        emps = cur.fetchall()
        if not emps:
            print("[demo] 无任何员工，无法继续生成其余演示数据")
            return
        emp_ids = [e["employee_id"] for e in emps]

        # 2) attendance（近 3 个月：当月 + 前 2 个月）
        cur.execute(f"SELECT COUNT(*) c FROM {DB_NAME}.attendance")
        if cur.fetchone()["c"] == 0:
            today = date.today()
            total = 0
            for (yy, mm) in _recent_months(today, months_back=2):
                last_day = today if (yy == today.year and mm == today.month) else _last_day_of(yy, mm)
                total += _gen_attendance_month(conn, cur, emp_ids, yy, mm, last_day)
            conn.commit()
            print(f"[demo] 已生成考勤 {total} 条（{len(emp_ids)} 人 × 近 3 个月）")
        else:
            print("[demo] attendance 已有数据，跳过")

        # 3) employee_salary（每员工一行）
        cur.execute(f"SELECT COUNT(*) c FROM {DB_NAME}.employee_salary")
        if cur.fetchone()["c"] == 0:
            ins = ("INSERT INTO employee_salary (no, employee_id, name, position, "
                   "base_salary, performance_rating, performance_bonus, allowance, gross_salary) "
                   "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)")
            for i, e in enumerate(emps):
                base = DEMO_SALARY_MAP.get(e["position"], DEMO_SALARY_DEFAULT)
                rating = round(random.uniform(0.85, 1.20), 2)
                bonus = round(min(base * 0.4, 10000), 2)
                allowance = 500.0
                gross = round(base + bonus + allowance, 2)
                cur.execute(ins, (i + 1, e["employee_id"], e["name"], e["position"],
                                  base, rating, bonus, allowance, gross))
            conn.commit()
            print(f"[demo] 已生成薪资 {len(emps)} 条")
        else:
            print("[demo] employee_salary 已有数据，跳过")

        # 4) reimbursement（随机抽取若干员工）
        cur.execute(f"SELECT COUNT(*) c FROM {DB_NAME}.reimbursement")
        if cur.fetchone()["c"] == 0:
            cats = ["差旅费", "餐饮费", "办公用品", "交通费", "培训费"]
            stats = ["submitted", "approving", "approved", "rejected"]
            ins = ("INSERT INTO reimbursement (employee_id, applicant_name, category, "
                   "amount, status, approver_id) VALUES (%s,%s,%s,%s,%s,%s)")
            n = 0
            for e in random.sample(emps, min(8, len(emps))):
                cur.execute(ins, (e["employee_id"], e["name"], random.choice(cats),
                                  round(random.uniform(120, 2200), 2), random.choice(stats), None))
                n += 1
            conn.commit()
            print(f"[demo] 已生成报销 {n} 条")
        else:
            print("[demo] reimbursement 已有数据，跳过")

        # 5) assessment_log（随机抽取若干员工）
        cur.execute(f"SELECT COUNT(*) c FROM {DB_NAME}.assessment_log")
        if cur.fetchone()["c"] == 0:
            poses = [e["position"] for e in emps]
            ins = "INSERT INTO assessment_log (employee_id, position_queried) VALUES (%s,%s)"
            n = 0
            for e in random.sample(emps, min(6, len(emps))):
                cur.execute(ins, (e["employee_id"], random.choice(poses)))
                n += 1
            conn.commit()
            print(f"[demo] 已生成考核日志 {n} 条")
        else:
            print("[demo] assessment_log 已有数据，跳过")


def _reset_all(conn):
    with conn.cursor() as cur:
        cur.execute("SET FOREIGN_KEY_CHECKS=0")
        for t in ("assessment_log", "reimbursement", "attendance", "employee_salary", "assessment_stats", "employees"):
            cur.execute(f"DELETE FROM {DB_NAME}.{t}")
            cur.execute(f"ALTER TABLE {DB_NAME}.{t} AUTO_INCREMENT=1")
        cur.execute("SET FOREIGN_KEY_CHECKS=1")
    conn.commit()
    print("[reset] 已清空 6 张表")


def gen_assessment_stats_programmatic(conn):
    """兜底生成：无种子 SQL 时，为每位员工生成 1 条考核成绩，保证一键数据完整。"""
    with conn.cursor() as cur:
        cur.execute(
            f"SELECT employee_id, position FROM {DB_NAME}.employees ORDER BY employee_id"
        )
        emps = cur.fetchall()
        if not emps:
            print("[seed] 无员工，跳过考核成绩兜底生成")
            return
        positions = [e["position"] for e in emps]
        ins = ("INSERT INTO assessment_stats "
               "(employee_id, original_position, target_position, score, test_time) "
               "VALUES (%s,%s,%s,%s,%s)")
        n = 0
        now = datetime.now()
        for e in emps:
            target = random.choice(positions)
            score = round(random.uniform(60, 98), 1)
            test_time = now - timedelta(days=random.randint(15, 220))
            cur.execute(ins, (e["employee_id"], e["position"], target, score, test_time))
            n += 1
        conn.commit()
        print(f"[seed] 已程序化生成考核成绩 {n} 条（兜底）")


def seed_assessment_stats(conn):
    """考核成绩：优先从 contracts/assessment_stats_seed.sql 导入（同事 advanced_db_init.sql 转换）。

    幂等：assessment_stats 非空则跳过；种子 SQL 缺失或导入为空时，程序化兜底生成，
    保证「一键生成数据」始终产出完整的考核成绩。
    """
    with conn.cursor() as cur:
        cur.execute(f"SELECT COUNT(*) c FROM {DB_NAME}.assessment_stats")
        if cur.fetchone()["c"] > 0:
            print("[seed] assessment_stats 已有数据，跳过")
            return
    imported = 0
    if ASSESSMENT_SEED_SQL.exists():
        sql = ASSESSMENT_SEED_SQL.read_text(encoding="utf-8")
        stmts, buf = [], ""
        for line in sql.splitlines():
            if line.strip().startswith("--"):
                continue
            buf += line + "\n"
            if line.strip().endswith(";"):
                stmts.append(buf.strip())
                buf = ""
        with conn.cursor() as cur:
            for s in stmts:
                cur.execute(s)
            cur.execute(f"SELECT COUNT(*) c FROM {DB_NAME}.assessment_stats")
            imported = cur.fetchone()["c"]
        conn.commit()
        print(f"[seed] 已导入考核成绩（{ASSESSMENT_SEED_SQL.name}，{imported} 条 INSERT）")
    if imported == 0:
        print("[seed] 种子 SQL 缺失/为空，改用程序化兜底生成考核成绩")
        gen_assessment_stats_programmatic(conn)


def seed_demo(conn, reset=False):
    if reset:
        _reset_all(conn)
    gen_demo_dataset(conn)
    seed_assessment_stats(conn)
    print("[seed-demo] 演示数据生成完成")


def main():
    conn = connect(**{k: v for k, v in MYSQL.items()})
    try:
        ensure_database(conn)
        if "--seed-demo" in sys.argv:
            seed_demo(conn, reset="--reset" in sys.argv)
        elif "--seed-employees" in sys.argv:
            seed_employees(conn)
            if SEED_ATTENDANCE:
                gen_attendance(conn)
        elif SEED_ATTENDANCE:
            gen_attendance(conn)
        with conn.cursor() as cur:
            for t in ("employees", "attendance", "reimbursement", "employee_salary", "assessment_log", "assessment_stats"):
                cur.execute(f"SELECT COUNT(*) c FROM {DB_NAME}.{t}")
                print(f"  表 {t}: {cur.fetchone()['c']} 行")
    finally:
        conn.close()


if __name__ == "__main__":
    main()
